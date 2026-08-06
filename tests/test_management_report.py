"""Tests for management-report generation."""

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from src.reporting.management_report import (
    clean_sheet_name,
    normalise_excel_value,
    write_dataframe,
)


def test_clean_sheet_name() -> None:
    """Invalid Excel worksheet characters should be removed."""

    result = clean_sheet_name(
        "Pipeline/Audit:Report?"
    )

    assert result == "Pipeline-Audit-Report"


def test_clean_sheet_name_length() -> None:
    """Excel worksheet names should not exceed 31 characters."""

    result = clean_sheet_name(
        "A" * 50
    )

    assert len(result) == 31


def test_normalise_excel_timestamp() -> None:
    """Pandas timestamps should become Python datetimes."""

    value = pd.Timestamp(
        "2026-08-06 12:00:00"
    )

    result = normalise_excel_value(
        value
    )

    assert result.year == 2026
    assert result.month == 8


def test_write_dataframe(
    tmp_path: Path,
) -> None:
    """A DataFrame should be written as an Excel table."""

    from openpyxl import Workbook

    workbook = Workbook()
    worksheet = workbook.active

    dataframe = pd.DataFrame(
        {
            "part_number": [
                "PN-001",
                "PN-002",
            ],
            "quantity": [
                5,
                10,
            ],
        }
    )

    next_row = write_dataframe(
        worksheet,
        dataframe,
        start_row=1,
        table_name="TestTable",
        header_colour="1F4E78",
        header_text_colour="FFFFFF",
    )

    output_path = (
        tmp_path
        / "test_report.xlsx"
    )

    workbook.save(
        output_path
    )

    loaded = load_workbook(
        output_path
    )

    loaded_sheet = loaded.active

    assert (
        loaded_sheet["A1"].value
        == "part_number"
    )

    assert (
        loaded_sheet["A2"].value
        == "PN-001"
    )

    assert next_row == 5