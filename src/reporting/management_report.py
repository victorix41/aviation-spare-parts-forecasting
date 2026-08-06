"""Excel management-report generation for aviation spare parts."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from src.dashboards.data_access import DashboardRepository


INVALID_EXCEL_CHARACTERS = {
    "/": "-",
    "\\": "-",
    "?": "",
    "*": "",
    "[": "(",
    "]": ")",
    ":": "-",
}


def clean_sheet_name(
    value: str,
) -> str:
    """Return a valid Excel worksheet name."""

    output = str(value)

    for old, new in INVALID_EXCEL_CHARACTERS.items():
        output = output.replace(old, new)

    return output[:31]


def normalise_excel_value(
    value: object,
) -> object:
    """Convert pandas values into Excel-compatible values."""

    if value is None:
        return None

    if isinstance(value, pd.Timestamp):
        if pd.isna(value):
            return None

        return value.to_pydatetime()

    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        pass

    if isinstance(value, (dict, list, tuple, set)):
        return str(value)

    return value


def apply_title(
    worksheet,
    *,
    title: str,
    subtitle: str,
    primary_colour: str,
    header_text_colour: str,
) -> None:
    """Apply a consistent report title."""

    worksheet.merge_cells(
        start_row=1,
        start_column=1,
        end_row=1,
        end_column=8,
    )

    title_cell = worksheet.cell(
        row=1,
        column=1,
        value=title,
    )

    title_cell.font = Font(
        bold=True,
        size=18,
        color=header_text_colour,
    )

    title_cell.fill = PatternFill(
        fill_type="solid",
        fgColor=primary_colour,
    )

    title_cell.alignment = Alignment(
        horizontal="left",
        vertical="center",
    )

    worksheet.row_dimensions[1].height = 28

    worksheet.merge_cells(
        start_row=2,
        start_column=1,
        end_row=2,
        end_column=8,
    )

    subtitle_cell = worksheet.cell(
        row=2,
        column=1,
        value=subtitle,
    )

    subtitle_cell.font = Font(
        italic=True,
        size=10,
    )


def write_dataframe(
    worksheet,
    dataframe: pd.DataFrame,
    *,
    start_row: int,
    table_name: str,
    header_colour: str,
    header_text_colour: str,
) -> int:
    """Write a DataFrame and return the next available row."""

    if dataframe.empty:
        worksheet.cell(
            row=start_row,
            column=1,
            value="No records available.",
        )

        return start_row + 2

    columns = [
        str(column)
        for column in dataframe.columns
    ]

    for column_index, column_name in enumerate(
        columns,
        start=1,
    ):
        cell = worksheet.cell(
            row=start_row,
            column=column_index,
            value=column_name,
        )

        cell.font = Font(
            bold=True,
            color=header_text_colour,
        )

        cell.fill = PatternFill(
            fill_type="solid",
            fgColor=header_colour,
        )

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    for row_index, values in enumerate(
        dataframe.itertuples(
            index=False,
            name=None,
        ),
        start=start_row + 1,
    ):
        for column_index, value in enumerate(
            values,
            start=1,
        ):
            cell = worksheet.cell(
                row=row_index,
                column=column_index,
                value=normalise_excel_value(
                    value
                ),
            )

            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )

    end_row = start_row + len(dataframe)
    end_column = len(columns)

    if end_column > 0:
        table_reference = (
            f"A{start_row}:"
            f"{get_column_letter(end_column)}"
            f"{end_row}"
        )

        excel_table = Table(
            displayName=table_name,
            ref=table_reference,
        )

        excel_table.tableStyleInfo = (
            TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
        )

        worksheet.add_table(
            excel_table
        )

    return end_row + 2


def autofit_columns(
    worksheet,
    *,
    maximum_width: int = 45,
) -> None:
    """Apply readable column widths."""

    for column_cells in worksheet.columns:
        maximum_length = 0

        column_letter = get_column_letter(
            column_cells[0].column
        )

        for cell in column_cells:
            value = cell.value

            if value is None:
                continue

            maximum_length = max(
                maximum_length,
                len(str(value)),
            )

        worksheet.column_dimensions[
            column_letter
        ].width = min(
            maximum_length + 2,
            maximum_width,
        )


def apply_currency_formats(
    worksheet,
    dataframe: pd.DataFrame,
    *,
    start_row: int,
) -> None:
    """Apply currency formatting to recognised columns."""

    currency_columns = {
        "inventory_value_usd",
        "procurement_value_usd",
        "unit_price_usd",
        "low_confidence_exposure_usd",
    }

    for column_index, column_name in enumerate(
        dataframe.columns,
        start=1,
    ):
        if str(column_name) not in currency_columns:
            continue

        for row_index in range(
            start_row + 1,
            start_row + len(dataframe) + 1,
        ):
            worksheet.cell(
                row=row_index,
                column=column_index,
            ).number_format = (
                '$#,##0.00'
            )


def create_metadata_frame(
    *,
    report_title: str,
    database_path: Path,
    latest_pipeline_run: pd.DataFrame,
) -> pd.DataFrame:
    """Create management-report metadata."""

    rows = [
        {
            "field": "Report title",
            "value": report_title,
        },
        {
            "field": "Generated at",
            "value": datetime.now().isoformat(
                timespec="seconds"
            ),
        },
        {
            "field": "Database",
            "value": str(database_path),
        },
        {
            "field": "Report type",
            "value": (
                "Read-only management decision support"
            ),
        },
        {
            "field": "Automatic purchasing",
            "value": "Not permitted",
        },
        {
            "field": "Human approval",
            "value": "Required",
        },
    ]

    if not latest_pipeline_run.empty:
        pipeline_row = latest_pipeline_run.iloc[0]

        rows.extend(
            [
                {
                    "field": "Pipeline run ID",
                    "value": pipeline_row[
                        "pipeline_run_id"
                    ],
                },
                {
                    "field": "Pipeline status",
                    "value": pipeline_row[
                        "overall_status"
                    ],
                },
                {
                    "field": "Pipeline completed",
                    "value": pipeline_row[
                        "completed_at"
                    ],
                },
            ]
        )

    return pd.DataFrame(rows)


def generate_management_report(
    *,
    repository: DashboardRepository,
    database_path: Path,
    output_path: Path,
    report_settings: dict[str, Any],
) -> Path:
    """Generate the complete Excel management report."""

    colours = report_settings[
        "colours"
    ]

    maximum_rows = int(
        report_settings[
            "maximum_detail_rows"
        ]
    )

    workbook = Workbook()

    workbook.remove(
        workbook.active
    )

    report_title = str(
        report_settings["title"]
    )

    report_subtitle = str(
        report_settings["subtitle"]
    )

    latest_pipeline_run = (
        repository.load_latest_pipeline_run()
    )

    executive_kpis = pd.DataFrame(
        [
            repository.load_executive_kpis()
        ]
    )

    risk_summary = (
        repository.load_risk_summary()
    )

    model_summary = (
        repository.load_model_summary()
    )

    procurement = (
        repository
        .load_procurement_dashboard_data()
        .head(maximum_rows)
    )

    finance = (
        repository
        .load_finance_exposure()
        .head(maximum_rows)
    )

    engineering = (
        repository
        .load_engineering_review_data()
        .head(maximum_rows)
    )

    operations = (
        repository
        .load_operations_readiness_data()
        .head(maximum_rows)
    )

    quality = (
        repository
        .load_quality_review_data()
        .head(maximum_rows)
    )

    pipeline_stages = (
        repository.load_latest_pipeline_stages()
    )

    recent_pipeline_runs = (
        repository.load_recent_pipeline_runs(
            limit=20
        )
    )

    metadata = create_metadata_frame(
        report_title=report_title,
        database_path=database_path,
        latest_pipeline_run=(
            latest_pipeline_run
        ),
    )

    report_sheets = [
        (
            "Executive Summary",
            [
                (
                    "Executive KPIs",
                    executive_kpis,
                    "ExecutiveKPIs",
                ),
                (
                    "Risk Summary",
                    risk_summary,
                    "RiskSummary",
                ),
                (
                    "Forecast Models",
                    model_summary,
                    "ForecastModels",
                ),
            ],
        ),
        (
            "Procurement Review",
            [
                (
                    "Procurement Review Queue",
                    procurement,
                    "ProcurementReview",
                ),
            ],
        ),
        (
            "Finance Exposure",
            [
                (
                    "Financial Exposure",
                    finance,
                    "FinanceExposure",
                ),
            ],
        ),
        (
            "Engineering Review",
            [
                (
                    "Engineering Review",
                    engineering,
                    "EngineeringReview",
                ),
            ],
        ),
        (
            "Operations Readiness",
            [
                (
                    "Operations Readiness",
                    operations,
                    "OperationsReadiness",
                ),
            ],
        ),
        (
            "Quality Assurance",
            [
                (
                    "Quality Review",
                    quality,
                    "QualityReview",
                ),
            ],
        ),
        (
            "Pipeline Audit",
            [
                (
                    "Latest Pipeline Stages",
                    pipeline_stages,
                    "PipelineStages",
                ),
                (
                    "Recent Pipeline Runs",
                    recent_pipeline_runs,
                    "PipelineRuns",
                ),
            ],
        ),
        (
            "Report Metadata",
            [
                (
                    "Report Metadata",
                    metadata,
                    "ReportMetadata",
                ),
            ],
        ),
    ]

    for sheet_name, sections in report_sheets:
        worksheet = workbook.create_sheet(
            title=clean_sheet_name(
                sheet_name
            )
        )

        apply_title(
            worksheet,
            title=report_title,
            subtitle=(
                f"{report_subtitle} — "
                f"{sheet_name}"
            ),
            primary_colour=colours[
                "primary"
            ],
            header_text_colour=colours[
                "header_text"
            ],
        )

        current_row = 4

        for (
            section_title,
            dataframe,
            table_name,
        ) in sections:
            worksheet.cell(
                row=current_row,
                column=1,
                value=section_title,
            ).font = Font(
                bold=True,
                size=13,
                color=colours[
                    "primary"
                ],
            )

            current_row += 1

            table_start_row = current_row

            current_row = write_dataframe(
                worksheet,
                dataframe,
                start_row=current_row,
                table_name=table_name,
                header_colour=colours[
                    "primary"
                ],
                header_text_colour=colours[
                    "header_text"
                ],
            )

            apply_currency_formats(
                worksheet,
                dataframe,
                start_row=table_start_row,
            )

        worksheet.freeze_panes = "A5"

        autofit_columns(
            worksheet
        )

    quality_sheet = workbook[
        "Quality Assurance"
    ]

    quality_sheet.conditional_formatting.add(
        "A1:Z1000",
        CellIsRule(
            operator="equal",
            formula=['"Failed"'],
            fill=PatternFill(
                fill_type="solid",
                fgColor=colours[
                    "critical"
                ],
            ),
        ),
    )

    output_path.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    workbook.save(
        output_path
    )

    return output_path